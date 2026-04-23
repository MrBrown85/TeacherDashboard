"""Generate fullvision-documentation-inventory.xlsx — a comprehensive index of
every documentation artifact in the repo and on GitHub."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROWS = [
    # (Section, Type, Path, Size, Audience, Description, Notes)

    # 1. TOP-LEVEL
    ("1. Top-level", "Markdown", "README.md", "227 lines", "Repo visitor", "Project overview, features, tech stack, getting started, full commands table, Netlify env-vars table, sw.js cache-bump reminder, project structure, architecture summary, privacy/compliance", "Absorbed SETUP_INSTRUCTIONS.md content in docs-cleanup (Apr 18 2026)"),
    ("1. Top-level", "Markdown", "CLAUDE.md", "38 lines", "AI agent / Developer", "Repo charter: current status, hard rules, pass references, and design-worktree context", "Present on main; AGENTS.md absent"),
    ("1. Top-level", "Text", "_headers", "653 B", "Tooling", "Netlify security headers: X-Frame-Options, CSP note, Referrer-Policy, HSTS, per-filetype cache rules", ""),
    ("1. Top-level", "JSON", "manifest.json", "1.2 KB", "Tooling", "PWA manifest: app name, icons, theme color, shortcuts", "Browser-facing"),
    ("1. Top-level", "JSON", "package.json", "658 B", "Tooling", "npm scripts (dev, test, test:watch, test:e2e, format, build), project metadata, devDependencies", ""),
    ("1. Top-level", "—", "LICENSE", "(missing)", "Repo visitor", "Would define usage terms", "GAP — no LICENSE file; repo metadata also shows licenseInfo: null"),
    ("1. Top-level", "—", ".env.example", "15 lines", "Developer", "Template env vars for local signed-in dev (`SUPABASE_URL`, `SUPABASE_KEY`)", "Used by scripts/dev-local.mjs"),
    ("1. Top-level", "—", "CONTRIBUTING.md", "(missing)", "Contributor", "Would describe contribution workflow, code style, PR process", "GAP"),
    ("1. Top-level", "—", "CHANGELOG.md", "(missing)", "Developer", "Would track release notes", "GAP"),
    ("1. Top-level", "—", "AGENTS.md", "(missing)", "AI agent", "Optional alternate agent-conventions file", "Low priority because CLAUDE.md exists"),
    ("1. Top-level", "—", "CODE_OF_CONDUCT.md / SECURITY.md", "(missing)", "Contributor", "Community / security disclosure docs", "GAP"),

    # 2. BUILD / CI / TOOLING CONFIG
    ("2. Build & CI", "TOML", "netlify.toml", "1.4 KB", "Tooling", "Build command, per-filetype cache headers (JS/CSS 1hr, HTML no-cache, JSON 1day), edge function routing, legacy URL redirects, reserved /student /parent portal routes", ""),
    ("2. Build & CI", "JavaScript", "vitest.config.js", "190 B", "Tooling", "Vitest config: test dir, setup files, globals, test file patterns", ""),
    ("2. Build & CI", "JavaScript", "playwright.config.js", "415 B", "Tooling", "Playwright E2E config: baseURL localhost:8347, headless, screenshot on failure, chromium, auto web server", ""),
    ("2. Build & CI", "YAML", ".github/workflows/ci.yml", "1.6 KB", "Tooling", "GitHub Actions: runs npm test + prettier --check on push/PR; narrow-to-changed-files logic", ""),
    ("2. Build & CI", "Shell", "scripts/build.sh", "17 lines", "Tooling", "Netlify build: copies public files to dist/", ""),

    # 3. ARCHITECTURE
    ("3. docs/ architecture", "Markdown", "docs/ARCHITECTURE.md", "447 lines", "Developer", "System design: directory structure, script load order, page routing, data flow, calculation engine, demo mode", "Needs periodic refresh when sync/write architecture changes"),

    # 4. BACKEND DESIGN / IMPLEMENTATION DOCS
    ("4. docs/backend-design", "Markdown", "docs/backend-design/HANDOFF.md", "342 lines", "Developer / AI agent", "Authoritative implementation log, current-state summary, discovered gaps, and activity history", "Refreshed Apr 21 2026 to remove stale rebuild-v2 / push-embargo guidance"),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/INSTRUCTIONS.md", "441 lines", "Developer / AI agent", "Design-level implementation instructions: scope, strings, data rules, UI directives", "Operational branch/cutover notes were refreshed Apr 21 2026"),
    ("4. docs/backend-design", "Markdown", "codex.md", "current", "Developer / AI agent", "Only active repo work list; consolidates open operational, platform, validation, deferred, and external items", "Canonical work queue after docs cleanup (Apr 22 2026)"),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/DECISIONS.md", "221 lines", "Developer", "Answer-by-answer decision log backing the rebuild scope", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/erd.md", "745 lines", "Developer", "Canonical ERD including Pass D amendment", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/write-paths.md", "1534 lines", "Developer", "Write-path RPC design and sequencing", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/read-paths.md", "650 lines", "Developer", "Read-path computation and surface design", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/auth-lifecycle.md", "604 lines", "Developer", "Auth, session, soft-delete, and restore lifecycle design", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/offline-sync.md", "103 lines", "Developer", "Offline queue model and sync lifecycle", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/spec-vs-ui-diff.md", "140 lines", "Developer", "Spec/UI reconciliation log", "Likely needs replay against post-merge UI"),
    ("4. docs/backend-design", "HTML", "docs/backend-design/decisions.html", "generated", "Developer / Stakeholder", "Rendered questionnaire / decision mirror", "Likely needs regeneration from DECISIONS.md"),
    ("4. docs/backend-design", "SQL", "docs/backend-design/schema.sql", "581 lines", "Developer", "Canonical schema mirror for the backend-design doc set", "Regenerated from live DB during reconciliation"),
    ("4. docs/backend-design", "SQL", "docs/backend-design/rls-policies.sql", "382 lines", "Developer", "RLS policy mirror for gradebook-prod", ""),
    ("4. docs/backend-design", "SQL", "docs/backend-design/read-paths.sql", "842 lines", "Developer", "Read-path RPC mirror", ""),
    ("4. docs/backend-design", "SQL", "docs/backend-design/write-paths.sql", "2047 lines", "Developer", "Write-path RPC mirror", ""),
    ("4. docs/backend-design", "SQL", "docs/backend-design/smoke-tests.sql", "1225 lines", "Developer", "psql-runnable smoke suite for live DB verification", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/smoke-tests.README.md", "137 lines", "Developer", "How to run and extend the smoke suite", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/smtp-setup.md", "146 lines", "Developer / Ops", "Runbook for custom SMTP setup and verification", ""),
    ("4. docs/backend-design", "Markdown", "docs/backend-design/DESIGN-SYSTEM.md", "current", "Developer / Designer", "CSS token and component-pattern inventory used to preserve the existing visual language", ""),

    # 5. COMPLIANCE / GOVERNANCE
    ("5. docs/ compliance", "Markdown", "docs/Privacy_Impact_Assessment.md", "243 lines", "Compliance", "PIA (March 22 2026): personal info inventory, collection purposes, data residency ca-central-1, RLS, idle timeout, no student accounts", "BC FOIPPA-aligned"),
    ("5. docs/ compliance", "Markdown", "docs/Data_Retention_Policy.md", "133 lines", "Compliance", "FOIPPA retention policy: what's kept vs deleted, retention periods, secure deletion on account wipe", ""),
    ("5. docs/ compliance", "Markdown", "docs/Breach_Notification_Procedure.md", "197 lines", "Compliance", "Breach definition, detection channels, response steps, 24-72hr notification timeline", ""),

    # 6. DIAGRAMS
    ("6. docs/diagrams/", "Markdown", "docs/diagrams/README.md", "35 lines", "Developer", "Guide to the 10 numbered diagrams, recommended reading order, color legend (green=Supabase, red=LS, yellow=conditional, blue=UI, purple=projection, orange=edge)", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/01-system-architecture.drawio", "10 KB", "Developer", "System containers: teacher desktop, mobile PWA, login, Supabase (Auth/Postgres/Realtime), Netlify edge, SW", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/02-database-schema.drawio", "14 KB", "Developer", "Entity relationships: course_data JSON blobs, teacher_config, auth users, RLS policies", "Pre-canonical-schema — likely stale"),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/03-frontend-module-map.drawio", "13 KB", "Developer", "JS module ownership: which file handles each page/tab", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/04-auth-and-routing.drawio", "10 KB", "Developer", "Login flow: Supabase Auth → portal redirect → app boot", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/05-hydration-on-login.drawio", "13 KB", "Developer", "Data loading: initAllCourses + initData RPC fan-out; notes April 3-18 data-invisible bug", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/06-write-path-map.drawio", "24 KB", "Developer", "Entity write paths color-coded by persistence (fully/partially/LS-only)", "Likely stale after v2 merge"),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/07-score-entry-paths.drawio", "10 KB", "Developer", "Desktop vs mobile scoring flow", "Likely stale after v2 merge"),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/08-observation-lifecycle.drawio", "12 KB", "Developer", "Full observation CRUD — template for unwired entities", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/09-term-report-flow.drawio", "11 KB", "Developer", "Questionnaire → term rating → progress report render → print", ""),
    ("6. docs/diagrams/", "Draw.io", "docs/diagrams/10-service-worker-cache.drawio", "10 KB", "Developer", "SW lifecycle, fetch strategy, cache busting, why PWA users need Unregister + hard reload", ""),

    # 7. USER-FLOW DIAGRAMS (Lucidchart)
    ("7. User-flow diagrams", "Markdown", "docs/lucidchart-user-flowcharts.md", "73 lines", "End user / Dev", "Two Mermaid flowcharts: first-time teacher path; recurring assignment/grade/report workflow", ""),
    ("7. User-flow diagrams", "Mermaid", "docs/lucidchart-first-time-report-flow.mmd", "971 B", "End user", "Flowchart: login → reports for first-time teacher", ""),
    ("7. User-flow diagrams", "Mermaid", "docs/lucidchart-teacher-assignment-comment-grade-flow.mmd", "825 B", "End user", "Flowchart: recurring workflow (assignments → grading → reports)", ""),
    ("7. User-flow diagrams", "Draw.io", "docs/lucidchart-user-flows.drawio", "20 KB", "Developer", "Same user flows in Draw.io format", ""),
    ("7. User-flow diagrams", "Draw.io", "docs/lucidchart-assignment-comment-grade-flow.drawio", "9 KB", "Developer", "Assignment workflow Draw.io", ""),

    # 8. SUPERPOWERS (feature plans + design specs)
    ("8. docs/superpowers/", "Markdown", "docs/superpowers/plans/2026-04-20-post-reconciliation-backlog.md", "(deleted)", "AI agent / Dev", "Former post-reconciliation backlog", "Historical reference only; unresolved items were consolidated into codex.md"),
    ("8. docs/superpowers/", "Markdown", "docs/superpowers/plans/2026-04-21-ui-v1-feature-gap.md", "(deleted)", "AI agent / Dev", "Former UI-v1 backlog and shipped Tier-A summary", "Historical reference only; unresolved items were consolidated into codex.md"),
    ("8. docs/superpowers/", "Markdown", "docs/superpowers/shipped/2026-04-20-database-wiring-reconciliation.md", "current", "AI agent / Dev", "Shipped reconciliation plan documenting the rebuild merge path", ""),

    # 9. INLINE CODE DOCS (file headers only)
    ("9. Inline code headers", "JSDoc", "shared/data.js", "~2800 lines", "Developer", "Cache-through pattern comment, Student/Assessment/Observation/TermRating typedefs, sync flow, fire-and-forget RPCs", "Entry point to the data layer"),
    ("9. Inline code headers", "JSDoc", "shared/calc.js", "~1500 lines", "Developer", "Proficiency calculation engine comment, Score typedef, 4 calc methods (mostRecent/highest/mode/decayingAvg), memoization caches, pointsToProf", ""),
    ("9. Inline code headers", "JSDoc", "shared/constants.js", "~800 lines", "Developer", "Course/LearningTag/LearningSection typedefs; curriculum structure; course config shape", ""),
    ("9. Inline code headers", "JSDoc", "shared/supabase.js", "~600 lines", "Developer", "Portal convention block (teacher/student/parent), auth strategy, dev mode bypass, credential handling, client init", ""),
    ("9. Inline code headers", "Comment", "teacher/app.html", "~50 lines", "Developer", "Meta tags, script load-order comment, curriculum_data lazy-load note, mobile auto-redirect", ""),
    ("9. Inline code headers", "Comment", "teacher/router.js", "~800 lines", "Developer", "Hash-based SPA router comment + route table; page lifecycle (init/destroy), query params", ""),
    ("9. Inline code headers", "Comment", "teacher/ui.js", "~900 lines", "Developer", "Shared UI components: dock, sidebar, modals; toast system; event delegation", ""),
    ("9. Inline code headers", "Comment", "teacher/page-dashboard.js", "~1200 lines", "Developer", "Page module IIFE pattern comment; dashboard cards, class overview", ""),
    ("9. Inline code headers", "Comment", "teacher/page-assignments.js", "~1400 lines", "Developer", "Assignment CRUD, scoring UI, rubric linking", ""),
    ("9. Inline code headers", "Comment", "teacher/page-student.js", "~1100 lines", "Developer", "Individual student profile: score timeline, sparklines, insights", ""),
    ("9. Inline code headers", "Comment", "teacher/page-gradebook.js", "~1300 lines", "Developer", "Spreadsheet-style gradebook; pinnable columns; view modes", ""),
    ("9. Inline code headers", "Comment", "teacher/page-observations.js", "~800 lines", "Developer", "Observation capture: quick notes, sentiment tagging", ""),
    ("9. Inline code headers", "Comment", "teacher/page-reports.js", "~1500 lines", "Developer", "Report builder: block selection, drag reorder, narrative generation", ""),
    ("9. Inline code headers", "Comment", "teacher/report-blocks.js", "~700 lines", "Developer", "Individual report block renderers (narrative, late policy, grade summary)", ""),
    ("9. Inline code headers", "Comment", "teacher/report-questionnaire.js", "~1300 lines", "Developer", "Term-questionnaire UI; narrative contenteditable; dim/trait/strength pickers", ""),
    ("9. Inline code headers", "Comment", "teacher/dash-class-manager.js", "~1800 lines", "Developer", "Class manager: student CRUD, roster import, curriculum editor, course create/delete", ""),
    ("9. Inline code headers", "Comment", "teacher/teams-import.js", "~900 lines", "Developer", "CSV/Excel roster import using SheetJS; Teams format parsing", ""),
    ("9. Inline code headers", "Comment", "teacher/assign-collab.js", "~235 lines", "Developer", "Collaboration panel: pairs/groups generation, drag-drop members", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/shell.js", "~1400 lines", "Developer", "Mobile shell + boot flow: auth check, tab routing, pull-to-refresh; offline fallback; data-action delegation", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/components.js", "~900 lines", "Developer", "iOS-style UI components: sheets, toasts, swipe gestures, native patterns", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/tab-students.js", "~1200 lines", "Developer", "Students tab: card stack, list, detail view, pull-to-refresh", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/tab-observe.js", "~900 lines", "Developer", "Observation feed + compose sheet (social feed pattern)", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/tab-grade.js", "~700 lines", "Developer", "Speed grader: one-student-at-a-time scoring optimized for mobile", "CRITICAL: still uses saveScores bulk (LS-only)"),
    ("9. Inline code headers", "Comment", "teacher-mobile/card-widget-editor.js", "~80 lines", "Developer", "Mobile card widget editor sheet: toggles + drag-reorder + reset", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/card-widgets.js", "~500 lines", "Developer", "Widget registry + renderers for the mobile student card", ""),
    ("9. Inline code headers", "Comment", "teacher-mobile/card-stack.js", "~400 lines", "Developer", "Swipeable student card stack component", ""),
    ("9. Inline code headers", "JSDoc", "netlify/edge-functions/inject-env.js", "49 lines", "Developer", "Edge function: env var injection, per-request CSP nonce, script/style tag rewriting", ""),
    ("9. Inline code headers", "Comment", "login-auth.js", "210 lines", "Developer", "Portal routing comment block (teacher/student/parent metadata), demo URL redirect (?demo=1), device-based mobile redirect", ""),
    ("9. Inline code headers", "Comment", "login.html", "88 lines", "End user", "Sign in / sign up tabs, error/success messages, demo mode button", ""),
    ("9. Inline code headers", "Comment", "sw.js", "~150 lines", "Developer", "Service worker: IMPORTANT CACHE_NAME comment, network-first strategy, offline fallback, PWA install", "Currently v34 on main; v35 pending"),

    # 10. TEST INFRASTRUCTURE

    ("10. Test infra", "JavaScript", "tests/setup.js", "86 lines", "Developer", "Vitest setup: browser globals shim (window, localStorage, DOM), script-tag loader for shared modules", ""),
    ("10. Test infra", "JavaScript", "e2e/helpers.js", "428 lines", "Developer", "Playwright helpers: auth seeding, fake Supabase client, localStorage fixtures", ""),
    ("10. Test infra", "—", "tests/README.md", "(missing)", "Developer", "Would document unit-test conventions, fake-client pattern, how to add tests", "GAP — 37 test files with no guide"),
    ("10. Test infra", "—", "e2e/README.md", "(missing)", "Developer", "Would document E2E setup, base URL, playwright.config expectations", "GAP — 16 E2E specs with no guide"),

    # 11. _dev/ INTERNAL DOCS
    ("11. _dev/ internal", "Markdown", "_dev/docs/Apple macOS Design Rulebook.md", "1291 lines", "Designer / AI", "Apple HIG reference for native-feel UI patterns", ""),
    ("11. _dev/ internal", "Markdown", "_dev/docs/TeacherDashboard Design Audit.md", "1645 lines", "Designer", "Visual design review notes; UI critique", ""),
    ("11. _dev/ internal", "Python", "_dev/build_curriculum_index.py", "(script)", "Developer", "Build: converts BC Curriculum Documents/*.json → indexed curriculum_data.js", ""),
    ("11. _dev/ internal", "Python", "_dev/serve.py", "(script)", "Developer", "Local dev server alternative to `npx serve`", ""),
    ("11. _dev/ internal", "Python", "_dev/make-inputs-xlsx.py", "(script)", "Developer", "Generator: fullvision-user-inputs.xlsx (template for this doc inventory script)", ""),
    ("11. _dev/ internal", "Python", "_dev/make-docs-xlsx.py", "(this script)", "Developer", "Generator: this documentation inventory xlsx", ""),
    ("11. _dev/ internal", "JSON", "_dev/BC Curriculum Documents/*.json", "300+ files", "Data", "BC Ministry of Education curriculum standards (Grades 8-12, all subjects)", "Source for curriculum_data.js"),

    # 12. GITHUB REPO METADATA (manual snapshot)
    ("12. GitHub metadata", "GitHub", "Repository", "MrBrown85/FullVision", "Repo visitor", "Public repo, default branch: main, visibility: public", ""),
    ("12. GitHub metadata", "GitHub", "Description", "(empty)", "Repo visitor", "No one-line repo description set", "GAP — add a description for GitHub search"),
    ("12. GitHub metadata", "GitHub", "Homepage URL", "(empty)", "Repo visitor", "No homepage linked (e.g. fullvision.ca)", "GAP — easy win"),
    ("12. GitHub metadata", "GitHub", "License", "(none)", "Repo visitor", "No LICENSE file; licenseInfo is null", "GAP"),
    ("12. GitHub metadata", "GitHub", "README (served)", "README.md", "Repo visitor", "Main README rendered on repo homepage", ""),
    ("12. GitHub metadata", "GitHub", "Wiki", "Enabled", "Repo visitor", "Wiki tab available but empty (no wiki pages created)", "Unused surface"),
    ("12. GitHub metadata", "GitHub", "Issues", "Enabled", "Contributor", "Issues tab open", ""),
    ("12. GitHub metadata", "GitHub", "Discussions", "Disabled", "Contributor", "Not enabled", ""),
    ("12. GitHub metadata", "GitHub", "Projects", "Enabled", "Contributor", "Projects enabled", ""),
    ("12. GitHub metadata", "GitHub", "GitHub Pages", "Not configured", "Repo visitor", "No Pages site (/pages endpoint returns 404)", ""),
    ("12. GitHub metadata", "GitHub", "Releases", "0", "Repo visitor", "No releases published", "GAP — no version history"),
    ("12. GitHub metadata", "GitHub", "Branches", "(refresh via gh api)", "Developer", "Branch list changes frequently; rerun the GitHub query when refreshing this workbook", "Local worktrees/branches were consolidated Apr 21 2026"),
    ("12. GitHub metadata", "File", ".github/workflows/ci.yml", "1.6 KB", "Tooling", "CI workflow (already listed in section 2)", ""),
    ("12. GitHub metadata", "—", ".github/PULL_REQUEST_TEMPLATE.md", "(missing)", "Contributor", "Would standardize PR descriptions", "GAP"),
    ("12. GitHub metadata", "—", ".github/ISSUE_TEMPLATE/", "(missing)", "Contributor", "Would standardize issue filing", "GAP"),
    ("12. GitHub metadata", "—", ".github/CODEOWNERS", "(missing)", "Contributor", "Would auto-request reviewers", "GAP"),
    ("12. GitHub metadata", "—", ".github/dependabot.yml", "(missing)", "Developer", "Would auto-update dependencies", "GAP"),
]

GAPS = [
    # (Area, Gap, Impact, Effort)
    ("Governance", "No LICENSE file", "Can't be legally reused; 'licenseInfo: null' on GitHub", "5 min — add MIT or similar"),
    ("Governance", "No CONTRIBUTING.md", "Contributors have to guess workflow and code style", "30 min"),
    ("Governance", "No CODE_OF_CONDUCT.md / SECURITY.md", "No reporting channel documented", "15 min each — use standard templates"),
    ("Governance", "No CHANGELOG.md", "No release notes; no version history for end-users or devs", "Start now; maintain per PR"),
    ("Governance", "Empty repo description and homepage URL", "GitHub search doesn't find it; no link to fullvision.ca", "2 min via gh repo edit"),
    ("AI / agents", "AGENTS.md absent", "Low impact because CLAUDE.md already documents agent conventions", "Optional"),
    ("Tests", "No tests/README.md or e2e/README.md", "37 unit specs + 16 E2E specs without contributor guide; test-writing conventions not documented", "45 min — explain fake Supabase client pattern, how to add a spec"),
    ("Tests", "Inline test comments sparse", "Complex test intent hard to follow", "Ongoing — add as you touch tests"),
    ("GitHub surface", "No PR / issue templates", "PRs vary in quality; no issue triage template", "15 min each"),
    ("GitHub surface", "No CODEOWNERS", "No auto-review assignment", "5 min"),
    ("GitHub surface", "No Dependabot config", "Deps drift silently", "5 min"),
    ("GitHub surface", "Wiki enabled but empty", "Surface area with no content", "Disable or populate"),
    ("Reference drift", "docs/ARCHITECTURE.md write/sync section can drift from the actual v2 queue and RPC dispatch paths", "Readers may infer the removed legacy retry bridge still exists", "20 min — audit whenever sync behavior changes"),
    ("Reference drift", "docs/diagrams/02-database-schema.drawio predates canonical schema", "Still shows course_data JSONB blobs; current DB uses academics/identity/canonical schemas", "1-2 hr — redraw"),
    ("Reference drift", "Documentation inventory workbook is manually generated from this script", "Workbook drifts whenever docs are added, renamed, or reclassified unless the script is updated and rerun", "15–30 min for each inventory refresh"),
    ("Feature coverage", "curriculum_data.js (994 KB) not documented", "Readers don't understand curriculum data model", "30 min — add section to ARCHITECTURE.md referencing _dev/build_curriculum_index.py"),
    ("Feature coverage", "Mobile platform-specific quirks undocumented", "iOS PWA cache / SW / home-screen install quirks live only in commit messages", "1 hr — write docs/MOBILE_PWA_NOTES.md"),
    ("Feature coverage", "Supabase migrations not in repo", "Only aggregated schema.sql; per-migration narrative is in Supabase dashboard only", "Export to supabase/migrations/ via CLI for future migrations"),
]

SUMMARY = [
    ("Category", "Count", "Notes"),
    ("Top-level docs (present)", 6, "README, CLAUDE.md, _headers, manifest.json, package.json, .env.example"),
    ("Build & CI configs", 5, "netlify.toml, vitest.config.js, playwright.config.js, ci.yml, build.sh"),
    ("Architecture + compliance docs", 4, "ARCHITECTURE, PIA, Retention, Breach (under-the-hood HTML removed)"),
    ("Backend-design docs", 19, "HANDOFF, INSTRUCTIONS, TASKS, DECISIONS, pass docs, SQL mirrors, smoke tests, SMTP runbook"),
    ("Diagrams (draw.io / mmd)", 15, "10 numbered + 1 diagram-README + 4 lucidchart user-flow files"),
    ("Superpowers plans + specs", 3, "2 active plans + 1 shipped reconciliation plan"),
    ("Inline code file headers", 26, "shared, teacher, teacher-mobile, netlify, login, sw.js"),
    ("Schema / DB docs", 6, "root schema mirror + backend-design schema/rls/read/write/smoke SQL docs"),
    ("Test infra docs", 2, "setup.js + helpers.js; no READMEs"),
    ("_dev/ internal docs + scripts", 7, "macOS Rulebook, Dashboard Audit, curriculum + server + xlsx scripts, BC curriculum JSONs"),
    ("GitHub metadata rows", 15, "repo settings + branches + missing templates"),
    ("Identified gaps", 18, "See Gaps sheet"),
]

OUT = "/Users/colinbrown/Documents/FullVision/fullvision-documentation-inventory.xlsx"

wb = Workbook()

# Sheet 1: All Docs
ws = wb.active
ws.title = "All Docs"

HEADERS = ["#", "Section", "Type", "Path", "Size", "Audience", "Description", "Notes"]
ws.append(HEADERS)

for i, r in enumerate(ROWS, 1):
    ws.append([i, r[0], r[1], r[2], r[3], r[4], r[5], r[6]])

# Styling
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=10)
wrap_align = Alignment(vertical="top", wrap_text=True)

fill_dev = PatternFill("solid", start_color="E2EFDA")      # green — developer
fill_user = PatternFill("solid", start_color="DDEBF7")     # blue — end user / teacher
fill_comp = PatternFill("solid", start_color="FFF2CC")     # yellow — compliance
fill_agent = PatternFill("solid", start_color="FCE4D6")    # orange — AI / contributor
fill_tool = PatternFill("solid", start_color="F2F2F2")     # grey — tooling / build
fill_missing = PatternFill("solid", start_color="FFCCCC")  # red-ish — missing

def audience_fill(audience, ftype, path):
    if "(missing)" in (path or "") or "(not in repo)" in (path or "") or "(not present" in (path or ""):
        return fill_missing
    a = (audience or "").lower()
    if "compliance" in a:
        return fill_comp
    if "ai" in a or "contributor" in a:
        return fill_agent
    if "end user" in a or "teacher" in a or "repo visitor" in a:
        return fill_user
    if "tooling" in a or "data" in a:
        return fill_tool
    return fill_dev

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx in range(2, ws.max_row + 1):
    path = ws.cell(row=row_idx, column=4).value
    audience = ws.cell(row=row_idx, column=6).value
    ftype = ws.cell(row=row_idx, column=3).value
    fill = audience_fill(audience, ftype, path)
    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.fill = fill

ws.freeze_panes = "A2"
widths = {"A": 5, "B": 22, "C": 12, "D": 52, "E": 18, "F": 20, "G": 55, "H": 38}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.auto_filter.ref = f"A1:H{ws.max_row}"

# Sheet 2: Gaps & Stale refs
gaps = wb.create_sheet("Gaps & Stale refs")
gaps.append(["#", "Area", "Gap / Stale reference", "Impact", "Effort to fix"])
for i, g in enumerate(GAPS, 1):
    gaps.append([i, g[0], g[1], g[2], g[3]])

for cell in gaps[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx in range(2, gaps.max_row + 1):
    for col_idx in range(1, 6):
        cell = gaps.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align

gaps.freeze_panes = "A2"
gaps.column_dimensions["A"].width = 5
gaps.column_dimensions["B"].width = 20
gaps.column_dimensions["C"].width = 60
gaps.column_dimensions["D"].width = 50
gaps.column_dimensions["E"].width = 35
gaps.auto_filter.ref = f"A1:E{gaps.max_row}"

# Sheet 3: Summary
summ = wb.create_sheet("Summary")
summ.append(SUMMARY[0])
for r in SUMMARY[1:]:
    summ.append(list(r))

total = sum(r[1] for r in SUMMARY[1:] if isinstance(r[1], int))
summ.append([])
summ.append(["Total rows on 'All Docs' sheet", len(ROWS), f"(Summary counts overlap — 'Gaps' are a separate view, not additional rows)"])

for cell in summ[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx in range(2, summ.max_row + 1):
    for col_idx in range(1, 4):
        cell = summ.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align

# Bold the final total
summ.cell(row=summ.max_row, column=1).font = Font(name="Arial", bold=True, size=11)
summ.cell(row=summ.max_row, column=2).font = Font(name="Arial", bold=True, size=11)

summ.column_dimensions["A"].width = 42
summ.column_dimensions["B"].width = 10
summ.column_dimensions["C"].width = 70

# Sheet 4: Legend
legend = wb.create_sheet("Legend")
legend.append(["Item", "Meaning"])
legend.append(["Audit date", "2026-04-21 (refreshed after main-branch reconciliation + docs sweep)"])
legend.append(["Local checkout", "/Users/colinbrown/Documents/FullVision"])
legend.append(["GitHub repo", "MrBrown85/FullVision (public, main branch)"])
legend.append([])
legend.append(["Audience color coding (All Docs sheet)", ""])
legend.append(["Green", "Developer / architecture"])
legend.append(["Blue", "End user / teacher-facing"])
legend.append(["Yellow", "Compliance / governance"])
legend.append(["Orange", "AI agent / contributor"])
legend.append(["Grey", "Tooling / build / data source"])
legend.append(["Red-pink", "Missing / absent (gap row)"])
legend.append([])
legend.append(["Source references", ""])
legend.append(["GitHub metadata queries", "gh repo view MrBrown85/FullVision --json ..."])
legend.append(["Branches query", "gh api repos/MrBrown85/FullVision/branches"])
legend.append(["Pages query", "gh api repos/MrBrown85/FullVision/pages (returns 404 → no Pages site)"])
legend.append(["Releases query", "gh api repos/MrBrown85/FullVision/releases (length: 0)"])
legend.append([])
legend.append(["Companion file", "fullvision-user-inputs.xlsx — 313 user inputs across 22 sections"])

for cell in legend[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx in range(2, legend.max_row + 1):
    for col_idx in range(1, 3):
        cell = legend.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align

legend.column_dimensions["A"].width = 38
legend.column_dimensions["B"].width = 80

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"  All Docs: {len(ROWS)} rows")
print(f"  Gaps & Stale refs: {len(GAPS)} rows")
print(f"  Summary: {len(SUMMARY)} rows")
print(f"  Sections in All Docs: {len(set(r[0] for r in ROWS))}")
