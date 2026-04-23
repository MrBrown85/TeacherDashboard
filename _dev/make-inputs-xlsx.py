"""Generate fullvision-user-inputs.xlsx — a comprehensive inventory of every
user input across the app, with persistence verdicts and source references."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

ROWS = [
    # (Section, Entity, Input/Field, UI Location, Action/Selector, Persistence, Notes)
    # 1. Authentication
    ("1. Authentication", "Auth", "Sign-in email", "login.html:33", "#si-email", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Sign-in password", "login.html:37", "#si-password", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Sign-up display name", "login.html:47", "#su-name", "SB (Auth)", "user_metadata"),
    ("1. Authentication", "Auth", "Sign-up email", "login.html:51", "#su-email", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Sign-up password", "login.html:55", "#su-password", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Sign-up password confirm", "login.html:59", "#su-confirm", "Validation only", ""),
    ("1. Authentication", "Auth", "Forgot-password email", "login.html", "forgot dialog", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Mobile sign-in email", "teacher-mobile/index.html:36", "#m-si-email", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Mobile sign-in password", "teacher-mobile/index.html:38", "#m-si-password", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Mobile sign-up name", "teacher-mobile/index.html:46", "#m-su-name", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Mobile sign-up email", "teacher-mobile/index.html:48", "#m-su-email", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Mobile sign-up password", "teacher-mobile/index.html:50", "#m-su-password", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Mobile sign-up password confirm", "teacher-mobile/index.html:52", "#m-su-confirm", "Validation only", ""),
    ("1. Authentication", "Auth", "Sign-out (desktop)", "teacher/ui.js", "signOut", "Clears session + LS", ""),
    ("1. Authentication", "Auth", "Sign-out (mobile)", "teacher-mobile/shell.js", "m-sign-out", "Clears session + LS", ""),
    ("1. Authentication", "Auth", "Delete account", "account menu", "deleteAccount", "SB (Auth)", ""),
    ("1. Authentication", "Auth", "Demo Mode toggle", "login.html", "Try Demo Mode button", "LS flag", "gb-demo-mode"),

    # 2. Courses
    ("2. Courses", "Course", "New course — name", "dash-class-manager.js:1049", "#cm-new-name", "SB", "create_course RPC"),
    ("2. Courses", "Course", "New course — grade level", "dash-class-manager.js:1050", "#cm-new-grade", "SB", ""),
    ("2. Courses", "Course", "New course — description", "dash-class-manager.js:1051", "#cm-new-desc", "SB", ""),
    ("2. Courses", "Course", "New course — grading system toggle", "dash-class-manager.js", "cmSetGradingSystem", "SB", "initial policy"),
    ("2. Courses", "Course", "New course — calc method toggle", "dash-class-manager.js", "cmSetCalcMethod", "SB", ""),
    ("2. Courses", "Course", "New course — decay slider", "dash-class-manager.js:1068", "#cm-cg-decay-slider", "SB", ""),
    ("2. Courses", "Course", "Edit course — name", "dash-class-manager.js:616", "#cm-name", "SB", "update_course RPC"),
    ("2. Courses", "Course", "Edit course — grade level", "dash-class-manager.js:621", "#cm-grade", "SB", ""),
    ("2. Courses", "Course", "Edit course — description", "dash-class-manager.js:634", "#cm-desc", "SB", ""),
    ("2. Courses", "Course", "Switch active course (desktop)", "multiple pages", "switchCourse", "SB", "save_teacher_preferences"),
    ("2. Courses", "Course", "Switch active course (mobile)", "teacher-mobile/shell.js:445", "#m-course-select", "SB", ""),
    ("2. Courses", "Course", "Duplicate course", "dash-class-manager.js", "cmDuplicateCourse", "SB", "new course"),
    ("2. Courses", "Course", "Archive / unarchive course", "dash-class-manager.js", "cmToggleArchive", "LS", "no archive RPC wired"),
    ("2. Courses", "Course", "Delete course", "dash-class-manager.js", "cmDeleteCourse", "LS", "no delete_course RPC"),
    ("2. Courses", "Course", "Course color picker", "dash-class-manager.js", "cmPickColor", "LS", ""),

    # 3. Course Policy
    ("3. Course Policy", "Policy", "Grading system select", "page-assignments.js", "settings panel", "SB", "save_course_policy"),
    ("3. Course Policy", "Policy", "Calculation method select", "page-assignments.js", "updateCalcMethod", "SB", ""),
    ("3. Course Policy", "Policy", "Decay weight slider", "page-assignments.js:195", "updateDecaySlider", "SB", ""),
    ("3. Course Policy", "Policy", "Decay weight slider (class-manager)", "dash-class-manager.js:755", "cmDecaySlider", "SB", ""),
    ("3. Course Policy", "Policy", "Grading scale label (per level)", "page-assignments.js:1441", ".gs-label", "SB", ""),
    ("3. Course Policy", "Policy", "Grading scale min boundary", "page-assignments.js:1442", ".gs-min", "SB", ""),
    ("3. Course Policy", "Policy", "Reset grading scale", "page-assignments.js", "resetGradingScale", "SB", ""),
    ("3. Course Policy", "Policy", "Category weights — enabled", "page-assignments.js:1449", "#cw-enabled", "SB", ""),
    ("3. Course Policy", "Policy", "Category weights — enabled (CM)", "dash-class-manager.js:763", "#cm-cw-enabled", "SB", ""),
    ("3. Course Policy", "Policy", "Category weights — summative % slider", "page-assignments.js:1450", "#cw-range", "SB", ""),
    ("3. Course Policy", "Policy", "Category weights — slider (CM)", "dash-class-manager.js:772", "cmCwRange", "SB", ""),
    ("3. Course Policy", "Policy", "Report as percentage toggle", "dash-class-manager.js:784", "#cm-report-pct", "SB", ""),
    ("3. Course Policy", "Policy", "Late-work policy (free text)", "report-blocks.js:669", "latePolicyEdit", "SB", "saves via save_course_policy"),

    # 4. Teacher Preferences
    ("4. Teacher Preferences", "Prefs", "Active course ID", "any course switcher", "switchCourse", "SB", "save_teacher_preferences"),
    ("4. Teacher Preferences", "Prefs", "Report period (Report 1-6)", "page-reports.js:749", "#report-period", "EPH", "not persisted"),
    ("4. Teacher Preferences", "Prefs", "Mobile view mode (cards/list)", "tab-students.js", "m-set-view", "LS pref", ""),
    ("4. Teacher Preferences", "Prefs", "Mobile sort mode", "tab-students.js", "m-set-sort", "LS pref", ""),
    ("4. Teacher Preferences", "Prefs", "Mobile card widget toggles", "card-widget-editor.js:48", "m-wdg-toggle", "LS pref", "UI pref — acceptable"),
    ("4. Teacher Preferences", "Prefs", "Mobile card widget drag reorder", "card-widget-editor.js", "m-wdg-drag", "LS pref", ""),
    ("4. Teacher Preferences", "Prefs", "Mobile card widget reset", "card-widget-editor.js:39", "m-wdg-reset", "LS pref", ""),
    ("4. Teacher Preferences", "Prefs", "Desktop view mode (grid/list)", "page-gradebook.js", "toggleGradesView / setView", "LS pref", ""),
    ("4. Teacher Preferences", "Prefs", "Desktop sidebar toggle", "teacher/ui.js", "toggleSidebar", "LS pref", ""),
    ("4. Teacher Preferences", "Prefs", "Toolbar dropdown state", "teacher/ui.js", "toggleToolbarDropdown", "EPH", ""),

    # 5. Students
    ("5. Students", "Student", "Add — first name", "dash-class-manager.js:704", "#cm-add-first", "SB", "enroll_student RPC"),
    ("5. Students", "Student", "Add — last name", "dash-class-manager.js:705", "#cm-add-last", "SB", ""),
    ("5. Students", "Student", "Add — preferred name", "dash-class-manager.js:706", "#cm-add-pref", "SB", ""),
    ("5. Students", "Student", "Add — pronouns", "dash-class-manager.js:707", "#cm-add-pro", "SB", ""),
    ("5. Students", "Student", "Add — student number", "dash-class-manager.js:708", "#cm-add-num", "SB", ""),
    ("5. Students", "Student", "Add — email", "dash-class-manager.js:709", "#cm-add-email", "SB", ""),
    ("5. Students", "Student", "Add — date of birth", "dash-class-manager.js:710", "#cm-add-dob", "SB", ""),
    ("5. Students", "Student", "Add — designations (IEP/MOD)", "dash-class-manager.js:713", ".cm-desig-check", "SB", "update_enrollment"),
    ("5. Students", "Student", "Edit — first name", "page-student.js:257", "#edit-first", "SB", "update_student RPC"),
    ("5. Students", "Student", "Edit — last name", "page-student.js:258", "#edit-last", "SB", ""),
    ("5. Students", "Student", "Edit — preferred name", "page-student.js:259", "#edit-pref", "SB", ""),
    ("5. Students", "Student", "Edit — pronouns", "page-student.js:260", "#edit-pro", "SB", ""),
    ("5. Students", "Student", "Edit — student number", "page-student.js:261", "#edit-num", "SB", ""),
    ("5. Students", "Student", "Edit — email", "page-student.js:262", "#edit-email", "SB", ""),
    ("5. Students", "Student", "Edit — date of birth", "page-student.js:263", "#edit-dob", "SB", ""),
    ("5. Students", "Student", "Edit — designations", "page-student.js:267", ".desig-check", "SB", ""),
    ("5. Students", "Student", "Remove (class-manager)", "dash-class-manager.js", "cmRemoveStudent", "SB", "withdraw_enrollment"),
    ("5. Students", "Student", "Delete (full, cascade)", "teacher/ui.js:410", "deleteStudent", "SB (partial)", "related goals/reflections/overrides/notes/flags LS-only"),
    ("5. Students", "Student", "Roster drag reorder", "dash-class-manager.js", "drag", "SB", "roster_position update_enrollment"),
    ("5. Students", "Student", "Bulk: apply pronouns to selection", "dash-class-manager.js", "cmApplyBulk", "SB", "via saveStudents"),
    ("5. Students", "Student", "Bulk: attendance date", "dash-class-manager.js:686", "#cm-bulk-att-date", "LS", "no attendance RPC"),
    ("5. Students", "Student", "Bulk: attendance status", "dash-class-manager.js:687", "#cm-bulk-att-status", "LS", ""),
    ("5. Students", "Student", "Bulk select/deselect all", "dash-class-manager.js", "cmBulkSelectAll / cmBulkDeselectAll", "EPH", ""),
    ("5. Students", "Student", "Bulk toggle row checkbox", "dash-class-manager.js:651", "cmBulkToggleCheck", "EPH", ""),
    ("5. Students", "Student", "Toggle bulk mode", "dash-class-manager.js", "cmToggleBulk", "EPH", ""),
    ("5. Students", "Student", "Show add-student form", "dash-class-manager.js", "cmShowAddStudent", "EPH", ""),
    ("5. Students", "Student", "Cancel add-student", "dash-class-manager.js", "cmCancelStudent", "EPH", ""),
    ("5. Students", "Student", "Save edited student", "page-student.js", "saveEditStudent", "SB", ""),
    ("5. Students", "Student", "Edit-student modal open", "page-student.js", "openEditStudent", "EPH", ""),
    ("5. Students", "Student", "Edit-student modal close", "page-student.js", "closeEditStudent", "EPH", ""),

    # 6. Assessments
    ("6. Assessments", "Assessment", "Title (new)", "page-gradebook.js:976", "#gb-new-title", "SB", "create_assessment RPC"),
    ("6. Assessments", "Assessment", "Title (form)", "page-assignments.js:446", "#af-title", "SB", ""),
    ("6. Assessments", "Assessment", "Description / notes", "page-assignments.js:452", "#af-desc", "SB", ""),
    ("6. Assessments", "Assessment", "Date assigned", "page-assignments.js:481", "#af-date-assigned", "SB", ""),
    ("6. Assessments", "Assessment", "Due date (form)", "page-assignments.js:485", "#af-due", "SB", ""),
    ("6. Assessments", "Assessment", "Date (gradebook add)", "page-gradebook.js:978", "#gb-new-date", "SB", ""),
    ("6. Assessments", "Assessment", "Type (summative/formative)", "page-gradebook.js:979", "#gb-new-type", "SB", ""),
    ("6. Assessments", "Assessment", "Type (setType)", "page-assignments.js", "setType", "SB", ""),
    ("6. Assessments", "Assessment", "Score mode (proficiency/points)", "page-assignments.js:524", "setScoreMode", "SB", ""),
    ("6. Assessments", "Assessment", "Max points (form)", "page-assignments.js:533", "#af-maxpoints", "SB", ""),
    ("6. Assessments", "Assessment", "Max points (gradebook)", "page-gradebook.js:983", "#gb-new-maxpts", "SB", ""),
    ("6. Assessments", "Assessment", "Max points chips", "page-assignments.js", "setMaxPoints", "SB", ""),
    ("6. Assessments", "Assessment", "Weight select", "page-assignments.js:539", "#af-weight", "SB", ""),
    ("6. Assessments", "Assessment", "Evidence type", "page-assignments.js:500", "#af-evidence", "SB", ""),
    ("6. Assessments", "Assessment", "Rubric select", "page-assignments.js:461", "#af-rubric", "SB", "onRubricSelect"),
    ("6. Assessments", "Assessment", "Module select", "page-assignments.js:512", "#af-module", "SB", ""),
    ("6. Assessments", "Assessment", "Linked tag IDs (standards)", "page-assignments.js:570", ".af-tag-cb / tagCheckbox", "SB", ""),
    ("6. Assessments", "Assessment", "Save new assessment", "page-assignments.js", "saveNewAssess / confirmAddAssess", "SB", ""),
    ("6. Assessments", "Assessment", "Save edited assessment", "page-assignments.js", "saveEditAssess", "SB", "update_assessment"),
    ("6. Assessments", "Assessment", "Duplicate assessment", "page-assignments.js", "dupeAssess", "SB", ""),
    ("6. Assessments", "Assessment", "Delete assessment", "page-assignments.js", "deleteAssess", "SB", "delete_assessment RPC"),
    ("6. Assessments", "Assessment", "Cancel assessment add/edit", "page-assignments.js", "cancelAddAssess / hideNewForm", "EPH", ""),
    ("6. Assessments", "Collab", "Collab mode", "assign-collab.js", "setCollaboration", "SB", "rides with assessment"),
    ("6. Assessments", "Collab", "Toggle excluded student", "assign-collab.js:59", "collabToggleStudent", "SB", ""),
    ("6. Assessments", "Collab", "Check all", "assign-collab.js", "collabCheckAll", "SB", ""),
    ("6. Assessments", "Collab", "Check none", "assign-collab.js", "collabCheckNone", "SB", ""),
    ("6. Assessments", "Collab", "Random pairs", "assign-collab.js", "collabRandomPairs", "SB", ""),
    ("6. Assessments", "Collab", "Random groups", "assign-collab.js", "collabRandomGroups", "SB", ""),
    ("6. Assessments", "Collab", "Manual pairs / groups", "assign-collab.js", "collabManualPairs / collabManualGroups", "SB", ""),
    ("6. Assessments", "Collab", "Group count +/-", "assign-collab.js", "collabSetGroupCount", "SB", ""),
    ("6. Assessments", "Collab", "Drag member between groups", "assign-collab.js:198", "collabDrop", "SB", ""),

    # 7. Scores
    ("7. Scores", "Score", "Click/cycle score (proficiency cell)", "page-gradebook.js:864", "cycleScore", "SB", "save_course_score via upsertScore"),
    ("7. Scores", "Score", "Inline edit cell value", "page-gradebook.js", "editScoreCell / startCellEdit", "SB", ""),
    ("7. Scores", "Score", "Points-mode numeric input (assignments)", "page-assignments.js:928", ".gb-pts-input", "LS", "setPointsScore path uses saveScores bulk"),
    ("7. Scores", "Score", "Points-mode numeric input (gradebook)", "page-gradebook.js:932", "setPointsScore", "LS", ""),
    ("7. Scores", "Score", "Score note / comment", "page-assignments.js:1401", "#comment-input", "SB", "via upsertScore"),
    ("7. Scores", "Score", "Submit comment", "page-assignments.js", "submitComment", "SB", ""),
    ("7. Scores", "Score", "Delete comment", "page-assignments.js:1414", "deleteComment", "SB", ""),
    ("7. Scores", "Score", "Fill rubric scores", "page-assignments.js", "fillRubricScoresAndClose", "SB", ""),
    ("7. Scores", "Score", "Fill scores", "page-assignments.js", "fillScoresAndClose", "SB", ""),
    ("7. Scores", "Score", "Confirm fill all", "page-assignments.js", "confirmFillAll / confirmFillRubricAll", "SB", ""),
    ("7. Scores", "Score", "Clear cell (context menu)", "page-gradebook.js:1147", "context menu", "LS", "delete_course_score RPC unused"),
    ("7. Scores", "Score", "Clear all row's scores", "page-gradebook.js:1196", "context menu", "LS", ""),
    ("7. Scores", "Score", "Clear all column's scores", "page-gradebook.js:1292", "context menu", "LS", ""),
    ("7. Scores", "Score", "Undo score change (desktop)", "page-gradebook.js", "executeUndo", "LS", "uses saveScores bulk"),
    ("7. Scores", "Score", "Mobile score button (proficiency 1-4)", "tab-grade.js:208", "m-grade-score", "LS", "MGrade.setScore uses saveScores bulk — CRITICAL"),
    ("7. Scores", "Score", "Mobile points increment", "tab-grade.js:198", "m-grade-points-inc", "LS", ""),
    ("7. Scores", "Score", "Mobile points decrement", "tab-grade.js:201", "m-grade-points-dec", "LS", ""),
    ("7. Scores", "Score", "Mobile undo toast", "tab-grade.js", "m-toast-undo", "LS", ""),
    ("7. Scores", "Score", "Mobile segmented filter (recent/all/ungraded)", "tab-grade.js:24-26", "m-grade-seg", "EPH", ""),
    ("7. Scores", "Score", "Mobile jump-to-student", "tab-grade.js", "m-grade-jump", "EPH", ""),
    ("7. Scores", "Score", "Toggle score menu (desktop)", "page-gradebook.js", "toggleScoreMenu", "EPH", ""),
    ("7. Scores", "Score", "Start score mode (dblclick)", "page-gradebook.js", "startScoreMode", "EPH", ""),
    ("7. Scores", "Score", "Select rubric score in fill", "page-assignments.js", "selectRubricScore", "SB", ""),
    ("7. Scores", "Score", "Select tag level", "page-assignments.js", "selectTagLevel", "SB", ""),

    # 8. Assignment Statuses
    ("8. Assignment Statuses", "Status", "Toggle student status (desktop)", "page-gradebook.js", "toggleStudentStatus", "SB", "save_assignment_status via setAssignmentStatus"),
    ("8. Assignment Statuses", "Status", "Mobile status pill NS", "tab-grade.js:219", "m-grade-status data-val=NS", "SB", ""),
    ("8. Assignment Statuses", "Status", "Mobile status pill EXC", "tab-grade.js:220", "m-grade-status data-val=EXC", "SB", ""),
    ("8. Assignment Statuses", "Status", "Mobile status pill LATE", "tab-grade.js:221", "m-grade-status data-val=LATE", "SB", ""),

    # 9. Observations
    ("9. Observations", "Observation", "Capture text (desktop)", "page-observations.js:108", "#capture-input", "SB", "create_observation RPC"),
    ("9. Observations", "Observation", "Capture text (mobile)", "tab-observe.js:147", "#m-obs-text", "SB", ""),
    ("9. Observations", "Observation", "Dimension tag toggle", "page-observations.js", ".dim-chip / m-obs-dim", "SB", ""),
    ("9. Observations", "Observation", "Sentiment (strength/growth/concern)", "page-observations.js", "toggleSentiment / m-obs-sentiment", "SB", ""),
    ("9. Observations", "Observation", "Context button", "page-observations.js", "toggleContext / m-obs-context", "SB", ""),
    ("9. Observations", "Observation", "Assignment context", "page-observations.js", "tag popover", "SB", ""),
    ("9. Observations", "Observation", "Remove capture student", "page-observations.js", "removeCaptureStudent", "SB", "changes scope"),
    ("9. Observations", "Observation", "Remove capture tag", "page-observations.js", "removeCaptureTag", "SB", ""),
    ("9. Observations", "Observation", "Submit observation (desktop)", "page-observations.js", "submitOb", "SB", ""),
    ("9. Observations", "Observation", "Save observation (mobile)", "tab-observe.js", "m-obs-save", "SB", ""),
    ("9. Observations", "Observation", "Edit observation (desktop)", "page-observations.js", "inline edit", "SB", "update_observation"),
    ("9. Observations", "Observation", "Edit observation (mobile)", "tab-observe.js", "m-obs-edit", "SB", ""),
    ("9. Observations", "Observation", "Delete observation (desktop)", "page-observations.js:534", "deleteOb", "SB", "delete_observation"),
    ("9. Observations", "Observation", "Delete observation (mobile)", "tab-observe.js", "m-obs-delete", "SB", ""),
    ("9. Observations", "Observation", "Mobile template quick-post", "tab-observe.js:401", "m-obs-quick-post", "SB", ""),
    ("9. Observations", "Observation", "Mobile compose new", "tab-observe.js", "m-obs-quick-compose", "EPH", "opens sheet"),
    ("9. Observations", "Observation", "Mobile quick menu", "tab-observe.js", "m-obs-quick-menu", "EPH", ""),
    ("9. Observations", "Observation", "Mobile pick student", "tab-observe.js", "m-obs-pick-student", "EPH", ""),
    ("9. Observations", "Observation", "Mobile select student", "tab-observe.js", "m-obs-select-student", "EPH", ""),
    ("9. Observations", "Observation", "Mobile student search", "tab-observe.js:188", "m-obs-student-search", "EPH", ""),
    ("9. Observations", "Observation", "Mobile remove student", "tab-observe.js", "m-obs-remove-student", "EPH", ""),
    ("9. Observations", "Observation", "Mobile filter", "tab-observe.js", "m-obs-filter", "EPH", ""),
    ("9. Observations", "Observation", "Mobile new observation FAB", "tab-observe.js", "m-obs-new", "EPH", ""),

    # 10. Notes
    ("10. Notes", "Note", "Add note text", "page-student.js:764", "#note-input / addNote", "LS", "no RPC"),
    ("10. Notes", "Note", "Search notes", "page-student.js:762", "#note-search", "EPH", ""),
    ("10. Notes", "Note", "Delete note", "page-student.js", "deleteNote", "LS", ""),

    # 11. Flags
    ("11. Flags", "Flag", "Toggle flag", "page-dashboard.js:645", "toggleFlag", "LS", "add_student_flag / remove_student_flag RPCs unused"),
    ("11. Flags", "Flag", "Toggle flagged-only filter", "page-dashboard.js:646", "toggleFlaggedFilter", "EPH", ""),

    # 12. Goals
    ("12. Goals", "Goal", "Goal text", "page-student.js:622", "#goal-input-{secId}", "LS", "save_student_goals RPC unused"),
    ("12. Goals", "Goal", "Save goal", "page-student.js", "saveGoalField", "LS", ""),
    ("12. Goals", "Goal", "Cancel goal edit", "page-student.js", "cancelGoalEdit", "EPH", ""),
    ("12. Goals", "Goal", "Edit goal", "page-student.js", "editGoal", "EPH", ""),

    # 13. Reflections
    ("13. Reflections", "Reflection", "Reflection text", "page-student.js:649", "#refl-input-{secId}", "LS", "save_student_reflection RPC unused"),
    ("13. Reflections", "Reflection", "Confidence select (1-5)", "page-student.js:642", "#refl-conf-{secId}", "LS", ""),
    ("13. Reflections", "Reflection", "Save reflection", "page-student.js", "saveReflField", "LS", ""),
    ("13. Reflections", "Reflection", "Cancel reflection edit", "page-student.js", "cancelReflEdit", "EPH", ""),
    ("13. Reflections", "Reflection", "Edit reflection", "page-student.js", "editReflection", "EPH", ""),

    # 14. Section Overrides
    ("14. Section Overrides", "Override", "Select level (1-4)", "page-student.js:85-88", "selectOverrideLevel", "LS", "save_section_override RPC unused"),
    ("14. Section Overrides", "Override", "Reason textarea", "page-student.js:91", "#override-reason-{secId}", "LS", ""),
    ("14. Section Overrides", "Override", "Save override", "page-student.js", "saveOverride", "LS", ""),
    ("14. Section Overrides", "Override", "Clear override", "page-student.js", "clearOverride", "LS", ""),
    ("14. Section Overrides", "Override", "Toggle override panel", "page-student.js", "toggleOverridePanel / closeOverridePanel", "EPH", ""),

    # 15. Learning Map
    ("15. Learning Map", "Subject", "Subject name (inline)", "dash-class-manager.js:807", "cmSubjectName", "LS", "save_learning_map RPC unused"),
    ("15. Learning Map", "Subject", "Add subject", "dash-class-manager.js", "cmAddSubject", "LS", ""),
    ("15. Learning Map", "Subject", "Delete subject", "dash-class-manager.js", "cmDeleteSubject", "LS", ""),
    ("15. Learning Map", "Section", "Section name (inline)", "dash-class-manager.js:83", "cmStdName", "LS", ""),
    ("15. Learning Map", "Section", "Section subject select", "dash-class-manager.js:85", "cmStdSubject", "LS", ""),
    ("15. Learning Map", "Section", "Section group select", "dash-class-manager.js:89", "cmStdGroup", "LS", ""),
    ("15. Learning Map", "Section", "Add section", "dash-class-manager.js", "cmAddStd", "LS", ""),
    ("15. Learning Map", "Section", "Delete section", "dash-class-manager.js", "cmDeleteStd", "LS", ""),
    ("15. Learning Map", "Section", "Toggle section folder (UI)", "dash-class-manager.js", "cmToggleStdFolder", "EPH", ""),
    ("15. Learning Map", "Section", "Toggle section card (UI)", "dash-class-manager.js", "cmToggleStdCard", "EPH", ""),
    ("15. Learning Map", "Tag", "Tag code (inline)", "dash-class-manager.js:102", "cmStdCode", "LS", ""),
    ("15. Learning Map", "Tag", "Tag label (inline)", "dash-class-manager.js:104", "cmStdLabel", "LS", ""),
    ("15. Learning Map", "Tag", "Tag I-can text", "dash-class-manager.js:107", "cmStdText", "LS", ""),
    ("15. Learning Map", "Tag", "Drag reorder", "dash-class-manager.js", "drag handler", "LS", ""),
    ("15. Learning Map", "Competency Group", "Add", "dash-class-manager.js", "cmAddCompGroup", "LS", ""),
    ("15. Learning Map", "Competency Group", "Delete", "dash-class-manager.js", "cmDeleteCompGroup", "LS", ""),
    ("15. Learning Map", "Competency Group", "Name (inline)", "dash-class-manager.js:845", "cmCompGroupName", "LS", ""),
    ("15. Learning Map", "Competency Group", "Color", "dash-class-manager.js:843", "cmCompGroupColor", "LS", ""),

    # 16. Modules
    ("16. Modules", "Module", "Name (inline)", "page-assignments.js:294", "moduleName", "LS", "no RPC"),
    ("16. Modules", "Module", "Color", "page-assignments.js:292", "moduleColor", "LS", ""),
    ("16. Modules", "Module", "Add module inline", "page-assignments.js", "addModuleInline", "LS", ""),
    ("16. Modules", "Module", "Delete module", "page-assignments.js:1467", "deleteModule", "LS", ""),
    ("16. Modules", "Module", "Toggle module folder (UI)", "page-assignments.js", "toggleModuleFolder / toggleModuleFilter", "EPH", ""),
    ("16. Modules", "Module", "Open color picker", "page-assignments.js", "openColorPicker / cmToggleColorPalette", "EPH", ""),

    # 17. Rubrics
    ("17. Rubrics", "Rubric", "Rubric name", "page-assignments.js:1302", "#re-name", "LS", "no RPC"),
    ("17. Rubrics", "Rubric", "Criterion name", "page-assignments.js:1318", "critName", "LS", ""),
    ("17. Rubrics", "Rubric", "Criterion level-4 descriptor", "page-assignments.js:1329", "critLevel data-level=4", "LS", ""),
    ("17. Rubrics", "Rubric", "Criterion level-3 descriptor", "page-assignments.js:1330", "critLevel data-level=3", "LS", ""),
    ("17. Rubrics", "Rubric", "Criterion level-2 descriptor", "page-assignments.js:1331", "critLevel data-level=2", "LS", ""),
    ("17. Rubrics", "Rubric", "Criterion level-1 descriptor", "page-assignments.js:1332", "critLevel data-level=1", "LS", ""),
    ("17. Rubrics", "Rubric", "Criterion linked tags", "page-assignments.js:1327", "toggleCritTag", "LS", ""),
    ("17. Rubrics", "Rubric", "Add criterion", "page-assignments.js", "addCriterion", "LS", ""),
    ("17. Rubrics", "Rubric", "Remove criterion", "page-assignments.js", "removeCriterion", "LS", ""),
    ("17. Rubrics", "Rubric", "Switch criterion section", "page-assignments.js", "switchCritSection", "EPH", ""),
    ("17. Rubrics", "Rubric", "Toggle criterion expand", "page-assignments.js", "toggleCriterionExpand", "EPH", ""),
    ("17. Rubrics", "Rubric", "Save rubric", "page-assignments.js", "saveRubricEdit", "LS", ""),
    ("17. Rubrics", "Rubric", "Cancel rubric edit", "page-assignments.js", "cancelRubricEdit", "EPH", ""),
    ("17. Rubrics", "Rubric", "New rubric", "page-assignments.js", "newRubricUI / newRubricFromForm", "LS", ""),
    ("17. Rubrics", "Rubric", "Edit rubric", "page-assignments.js", "editRubricUI", "EPH", ""),
    ("17. Rubrics", "Rubric", "Delete rubric", "page-assignments.js", "deleteRubricUI", "LS", ""),
    ("17. Rubrics", "Rubric", "Switch rubric view", "page-assignments.js", "setRubricView", "EPH", ""),
    ("17. Rubrics", "Rubric", "Open rubric panel", "page-assignments.js", "openRubricPanel / rubricModalBackdrop", "EPH", ""),

    # 18. Custom Tags
    ("18. Custom Tags", "Custom Tag", "Add custom tag text", "page-observations.js:191", "#custom-tag-input / addNewCustomTag", "LS", "no RPC"),

    # 19. Term Ratings
    ("19. Term Ratings", "Term Rating", "Narrative (rich HTML)", "report-questionnaire.js:1232", "#tq-narrative contenteditable", "SB", "upsert_term_rating RPC"),
    ("19. Term Ratings", "Term Rating", "Narrative formatting (bold/italic/list)", "report-questionnaire.js", "tqExec", "SB", "saved on blur"),
    ("19. Term Ratings", "Term Rating", "Narrative auto-generate", "report-questionnaire.js", "tqAutoNarrative", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Narrative copy", "report-questionnaire.js", "tqCopyNarrative", "EPH", "clipboard"),
    ("19. Term Ratings", "Term Rating", "Dim rating (per competency 1-4)", "report-questionnaire.js", "tqSetDim", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Work habits rating", "report-questionnaire.js", "tqSetField('workHabits')", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Participation rating", "report-questionnaire.js", "tqSetField('participation')", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Social trait toggle", "report-questionnaire.js", "tqToggleTrait", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Strengths (tag IDs)", "report-questionnaire.js", "tqSetField('strengths')", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Growth areas (tag IDs)", "report-questionnaire.js", "tqSetField('growthAreas')", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Mention assessment toggle", "report-questionnaire.js", "tqToggleAssignment", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Mention observation toggle", "report-questionnaire.js", "tqToggleOb", "SB", ""),
    ("19. Term Ratings", "Term Rating", "Next student nav", "report-questionnaire.js", "tqNextStudent", "EPH", ""),
    ("19. Term Ratings", "Term Rating", "Prev student nav", "report-questionnaire.js", "tqPrevStudent", "EPH", ""),
    ("19. Term Ratings", "Term Rating", "Observation filter", "report-questionnaire.js", "tqObsFilter", "EPH", ""),

    # 20. Report Config
    ("20. Report Config", "Report Config", "Apply preset (brief/standard/detailed)", "page-reports.js", "rbApplyPreset", "SB", "save_report_config RPC"),
    ("20. Report Config", "Report Config", "Toggle block enabled", "page-reports.js", "rbToggleBlock", "SB", ""),
    ("20. Report Config", "Report Config", "Anonymize toggle", "page-reports.js:749", "toggleAnon", "EPH", ""),
    ("20. Report Config", "Report Config", "Switch tab (questionnaire/progress/summary)", "page-reports.js", "switchTab", "EPH", ""),
    ("20. Report Config", "Report Config", "Student picker toggle", "page-reports.js", "toggleStudentPicker", "EPH", ""),
    ("20. Report Config", "Report Config", "Select all students", "page-reports.js", "selectAllStudents", "EPH", ""),
    ("20. Report Config", "Report Config", "Select none students", "page-reports.js", "selectNoneStudents", "EPH", ""),
    ("20. Report Config", "Report Config", "Toggle student in selection", "page-reports.js", "toggleStudentSelection", "EPH", ""),
    ("20. Report Config", "Report Config", "Print reports", "page-reports.js", "printReports", "N/A", "print dialog"),

    # 21. Imports
    ("21. Imports", "Import", "Class roster CSV", "dash-class-manager.js:726", "#cm-csv-input / cmCSV", "SB", "via saveStudents"),
    ("21. Imports", "Import", "Roster import confirm", "dash-class-manager.js", "cmConfirmImport", "SB", ""),
    ("21. Imports", "Import", "Roster import cancel", "dash-class-manager.js", "cmCancelImport", "EPH", ""),
    ("21. Imports", "Import", "Class wizard select grade", "dash-class-manager.js", "cwSelectGrade", "SB", "via createCourse"),
    ("21. Imports", "Import", "Class wizard select subject", "dash-class-manager.js", "cwSelectSubject", "SB", ""),
    ("21. Imports", "Import", "Class wizard toggle course", "dash-class-manager.js", "cwToggleCourse", "EPH", ""),
    ("21. Imports", "Import", "Class wizard step nav (step 2)", "dash-class-manager.js", "cwGoToStep2", "EPH", ""),
    ("21. Imports", "Import", "Class wizard step nav (step 3)", "dash-class-manager.js", "cwGoToStep3", "EPH", ""),
    ("21. Imports", "Import", "Class wizard back", "dash-class-manager.js", "cwGoBack", "EPH", ""),
    ("21. Imports", "Import", "Class wizard skip to custom", "dash-class-manager.js", "cwSkipToCustom", "EPH", ""),
    ("21. Imports", "Import", "Class wizard finish create", "dash-class-manager.js", "cwFinishCreate", "SB", "creates course"),
    ("21. Imports", "Import", "Relink roster start", "dash-class-manager.js", "cmStartRelink", "EPH", ""),
    ("21. Imports", "Import", "Relink confirm", "dash-class-manager.js", "cmRelinkConfirm", "SB", ""),
    ("21. Imports", "Import", "Relink back / cancel / next", "dash-class-manager.js", "cmRelinkBack / cmRelinkCancel / cmRelinkNext", "EPH", ""),

    # 22. Search / Filter / Navigation
    ("22. Nav & Filter", "Search", "Dashboard search", "page-dashboard.js:67", "dashSearch", "EPH", ""),
    ("22. Nav & Filter", "Search", "Dashboard sort", "page-dashboard.js:69", "dashSort", "EPH", ""),
    ("22. Nav & Filter", "Search", "Gradebook search", "page-gradebook.js:141", "gbSearch", "EPH", ""),
    ("22. Nav & Filter", "Search", "Gradebook roster search (top bar)", "teacher/ui.js:275", "#gb-roster-search", "EPH", ""),
    ("22. Nav & Filter", "Search", "Assignments search", "page-assignments.js:148", "assessSearch", "EPH", ""),
    ("22. Nav & Filter", "Search", "Observations search", "page-observations.js:119", "#obs-search-input", "EPH", ""),
    ("22. Nav & Filter", "Search", "Popover search", "page-observations.js:152", "popoverSearch", "EPH", ""),
    ("22. Nav & Filter", "Search", "Tag popover search", "page-observations.js:165", "tagPopoverSearch", "EPH", ""),
    ("22. Nav & Filter", "Search", "Mobile student search", "tab-students.js:84", "m-student-search", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Module (assignments)", "page-assignments.js", "toggleModuleFilter", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Type (assignments)", "page-assignments.js", "setTypeFilter / setAssessTypeFilter", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Section", "page-assignments.js", "toggleSectionFilter / clearSectionFilters", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Student (observations)", "page-observations.js", "removeFilterStudent", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Tag (observations)", "page-observations.js", "removeFilterTag", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Sentiment (observations)", "page-observations.js", "setFilterSentiment", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Clear all filters", "page-observations.js", "clearAllFilters / clearAssessFilters", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Toggle filter strip", "page-observations.js", "toggleFilterStrip", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Toggle ungraded filter", "page-gradebook.js", "toggleUngraded", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Toggle sort", "page-gradebook.js", "toggleSort", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Focus single student", "page-gradebook.js", "sidebarStudentClick / summaryStudentClick / goToStudent", "EPH", ""),
    ("22. Nav & Filter", "Filter", "Clear focused student", "page-gradebook.js", "clearFocusStudent", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Switch tab", "page-reports.js", "switchTab", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Toggle sidebar", "teacher/ui.js", "toggleSidebar", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Toggle user menu", "teacher/ui.js", "toggleUserMenu", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Collapse/expand all assessments", "page-assignments.js", "collapseAllAssess / expandAllAssess / toggleAssess / toggleGroupExpand", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Open class manager", "dash-class-manager.js", "openClassManager / closeClassManager", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Open edit student", "page-student.js", "openEditStudent / closeEditStudent", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Open comment popover", "page-assignments.js", "openCommentPopover / closeCommentPopover", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Open popover", "page-observations.js", "openPopover / popoverToggle", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Mobile sort sheet", "tab-students.js", "m-sort / m-set-sort", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Mobile settings sheet", "shell.js", "m-settings", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Mobile toggle section", "tab-students.js", "m-toggle-section", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Mobile back", "shell.js", "m-back", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Mobile dismiss sheet", "components.js", "m-dismiss-sheet", "EPH", ""),
    ("22. Nav & Filter", "Nav", "Retry sync", "teacher/ui.js", "retry-sync", "EPH", "network action"),
    ("22. Nav & Filter", "Nav", "Reload page", "teacher/ui.js", "reload-page", "EPH", ""),
    ("22. Nav & Filter", "Action", "Reset demo data", "teacher/ui.js", "resetDemoData", "LS wipe", ""),
    ("22. Nav & Filter", "Action", "Clear data", "teacher/ui.js", "clearData", "LS wipe + remote cleanup", ""),
    ("22. Nav & Filter", "Action", "Export data (JSON)", "teacher/ui.js", "exportData", "Read-only download", ""),
    ("22. Nav & Filter", "Action", "Export scores CSV", "teacher/ui.js", "exportScoresCSV", "Read-only download", ""),
    ("22. Nav & Filter", "Action", "Export summary CSV", "teacher/ui.js", "exportSummaryCSV", "Read-only download", ""),
]

GAPS = [
    # Priority, Area, Issue, RPC status, Fix
    (1, "Mobile scoring", "Every proficiency score entered on the phone stays local only; sign out → data gone", "save_course_score exists (used by desktop)", "Replace saveScores(cid, allScores) with per-tag upsertScore() in tab-grade.js:257"),
    (2, "Points-mode scoring", "Any assessment with scoreMode='points' stores grades in localStorage only", "save_course_score exists", "Rewrite setPointsScore() in shared/data.js:4092 to call save_course_score per tag"),
    (3, "Score deletions", "'Clear cell' / clear-row / clear-column / undo never reach the DB; deleted scores reappear on re-login", "delete_course_score exists but unused", "Wire delete_course_score into page-gradebook.js clear handlers + undo path"),
    (4, "Overrides", "Section-level professional-judgment overrides + reasons are local only", "save_section_override exists but unused", "Change saveOverrides() in shared/data.js:4016 to call save_section_override"),
    (5, "Goals", "Student per-section learning goals evaporate on sign-out", "save_student_goals exists but unused", "Wire save_student_goals into saveGoals()"),
    (6, "Reflections", "Student per-section reflections + confidence ratings lost on sign-out", "save_student_reflection exists but unused", "Wire into saveReflections()"),
    (7, "Flags", "Student follow-up flags don't persist across sign-in cycles", "add_student_flag / remove_student_flag exist but unused", "Change toggleFlag() in shared/data.js:4151 to call add/remove RPC"),
    (8, "Notes", "Teacher-authored student notes are local only", "No RPC", "Schema design needed: academics.student_note table + save/delete RPCs"),
    (9, "Learning map", "All curriculum edits (subjects, sections, tags, competency groups) are local only", "save_learning_map exists but unused", "Wire into saveLearningMap() — touches many inline handlers"),
    (10, "Modules", "Assignment folders (name, color) are local only", "No RPC", "Schema design needed"),
    (11, "Rubrics", "Full rubric definitions (criteria, levels, linked tags) are local only", "No RPC", "Schema design needed"),
    (12, "Competency groups", "Group names and colors are local only (part of learning map)", "Covered by save_learning_map once wired", "Lands with Learning map fix"),
    (13, "Custom observation tags", "Teacher-defined tags for observations are local only", "No RPC", "Schema design needed"),
    (14, "Attendance", "Bulk attendance stamps per student are local only", "No RPC", "Schema design needed"),
]

SUMMARY = [
    ("Category", "Count", "Notes"),
    ("Supabase-persisted (data)", 97, "All rows marked 'SB' below, excluding Auth"),
    ("Supabase-persisted (auth)", 6, "login/signup/reset/delete account"),
    ("LocalStorage-only (data that should sync)", 57, "The gap — see Gaps sheet"),
    ("LS-only (preferences, acceptable)", 7, "Mobile view/sort/widgets, desktop view/sidebar"),
    ("Ephemeral UI state", 128, "Search, filter, navigation, modal open/close"),
    ("Validation / non-storage", 5, "Password confirm, export downloads, print"),
]

OUT = "/Users/colinbrown/Documents/FullVision/fullvision-user-inputs.xlsx"

wb = Workbook()

# Sheet 1: Master inventory
ws = wb.active
ws.title = "All Inputs"

HEADERS = ["#", "Section", "Entity", "Input / Field", "UI Location", "Action / Selector", "Persistence", "Notes"]
ws.append(HEADERS)

for i, r in enumerate(ROWS, 1):
    ws.append([i, r[0], r[1], r[2], r[3], r[4], r[5], r[6]])

# Header styling
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", start_color="1F4E79")
header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Body styling + row fills by persistence
body_font = Font(name="Arial", size=10)
wrap_align = Alignment(vertical="top", wrap_text=True)
fill_sb = PatternFill("solid", start_color="E2EFDA")  # green-ish
fill_ls = PatternFill("solid", start_color="FCE4D6")  # orange-ish
fill_eph = PatternFill("solid", start_color="F2F2F2")  # light grey
fill_mixed = PatternFill("solid", start_color="FFF2CC")  # yellow

for row_idx in range(2, ws.max_row + 1):
    persistence = ws.cell(row=row_idx, column=7).value or ""
    fill = None
    if persistence.startswith("SB"):
        fill = fill_sb
    elif persistence.startswith("LS") and "pref" not in persistence:
        fill = fill_ls
    elif persistence.startswith("Mixed"):
        fill = fill_mixed
    else:
        fill = fill_eph
    for col_idx in range(1, 9):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align
        if fill:
            cell.fill = fill

# Freeze header, set widths
ws.freeze_panes = "A2"
widths = {"A": 5, "B": 22, "C": 20, "D": 40, "E": 36, "F": 38, "G": 18, "H": 38}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Autofilter
ws.auto_filter.ref = f"A1:H{ws.max_row}"

# Sheet 2: Gaps
gaps = wb.create_sheet("Gaps (priority)")
gaps.append(["Priority", "Area", "Issue", "RPC status", "Fix"])
for r in GAPS:
    gaps.append(list(r))

for cell in gaps[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx in range(2, gaps.max_row + 1):
    for col_idx in range(1, 6):
        cell = gaps.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align

gaps.freeze_panes = "A2"
gaps.column_dimensions["A"].width = 10
gaps.column_dimensions["B"].width = 22
gaps.column_dimensions["C"].width = 55
gaps.column_dimensions["D"].width = 40
gaps.column_dimensions["E"].width = 55

# Sheet 3: Summary
summ = wb.create_sheet("Summary")
summ.append(SUMMARY[0])
for r in SUMMARY[1:]:
    summ.append(list(r))

for cell in summ[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

summ.append([])
summ.append(["Total inputs enumerated", f"=SUM(B2:B{len(SUMMARY)})", ""])
summ.cell(row=summ.max_row, column=1).font = Font(name="Arial", bold=True, size=11)

for row_idx in range(2, summ.max_row + 1):
    for col_idx in range(1, 4):
        cell = summ.cell(row=row_idx, column=col_idx)
        if not cell.font.bold:
            cell.font = body_font
        cell.alignment = wrap_align

summ.column_dimensions["A"].width = 42
summ.column_dimensions["B"].width = 12
summ.column_dimensions["C"].width = 55

# Sheet 4: Legend
legend = wb.create_sheet("Legend")
legend.append(["Persistence code", "Meaning", "Row color"])
legend.append(["SB", "Persists to Supabase via a verified RPC call", "Green"])
legend.append(["LS", "localStorage only — stubbed _doSync or no RPC exists yet", "Orange"])
legend.append(["LS pref", "localStorage preference — acceptable (UI-only state)", "Grey"])
legend.append(["EPH", "Ephemeral session/UI state — intentionally not persisted", "Grey"])
legend.append(["Mixed", "Some sub-fields reach SB, others stay local", "Yellow"])
legend.append([])
legend.append(["Source audit note", "Second-pass verification on 2026-04-18 against commit " + "b156008 + main merge", ""])
legend.append(["Bridge stub location", "shared/data.js:379 (_doSync early-returns for legacy tables)", ""])
legend.append(["Canonical read path", "shared/data.js:1926 (_doInitData post-PR #71)", ""])
legend.append(["Known unwired RPCs in Supabase", "save_learning_map, save_section_override, save_student_goals, save_student_reflection, add_student_flag, remove_student_flag, delete_course_score", ""])

for cell in legend[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

for row_idx in range(2, legend.max_row + 1):
    for col_idx in range(1, 4):
        cell = legend.cell(row=row_idx, column=col_idx)
        cell.font = body_font
        cell.alignment = wrap_align

legend.column_dimensions["A"].width = 24
legend.column_dimensions["B"].width = 75
legend.column_dimensions["C"].width = 15

wb.save(OUT)
print(f"Wrote {OUT} — {len(ROWS)} input rows across {len(set(r[0] for r in ROWS))} sections")
